from flask import Flask, render_template, jsonify, request, session, redirect, url_for
import requests
from collections import defaultdict
from datetime import datetime, timedelta
import json
import os
import logging
import mimetypes
import re
import uuid
import secrets
from concurrent.futures import ThreadPoolExecutor, as_completed
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from functools import wraps
from email_service import (send_course_assignment_email, send_deadline_reminder_email,
                           send_course_removal_email, send_course_completion_email)
import assignment_analytics as aa
import db
import openpyxl
from io import BytesIO
import threading
import time
from dotenv import load_dotenv
from logging.handlers import RotatingFileHandler

load_dotenv()

# On Windows, mimetypes seeds itself from the registry, where HKCR\.js is often
# clobbered to text/plain by other installers. Browsers then refuse to execute
# our static JS under strict MIME checking, so pin the correct types here.
mimetypes.add_type('text/javascript', '.js')
mimetypes.add_type('text/css', '.css')
mimetypes.add_type('image/svg+xml', '.svg')
mimetypes.add_type('application/json', '.json')

# Configure logging — console + persistent rotating file (audit trail).
# Attach handlers directly to the root logger and clear any pre-existing ones
# (e.g. email_service.py calls logging.basicConfig() at import time, which would
# otherwise make our basicConfig a no-op and silently drop the file handler).
os.makedirs('logs', exist_ok=True)
_log_format = logging.Formatter('%(asctime)s %(levelname)s %(name)s: %(message)s')

_file_handler = RotatingFileHandler('logs/app.log', maxBytes=1_000_000, backupCount=5, encoding='utf-8')
_file_handler.setFormatter(_log_format)
_file_handler.setLevel(logging.INFO)

_console_handler = logging.StreamHandler()
_console_handler.setFormatter(_log_format)

_root_logger = logging.getLogger()
_root_logger.setLevel(logging.INFO)
for _h in list(_root_logger.handlers):
    _root_logger.removeHandler(_h)
_root_logger.addHandler(_console_handler)
_root_logger.addHandler(_file_handler)

logger = logging.getLogger(__name__)

app = Flask(__name__)
app.secret_key = os.environ['SECRET_KEY']

# Session security
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = os.environ.get('HTTPS_ONLY', 'false').lower() == 'true'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=8)
app.config['MAX_CONTENT_LENGTH'] = 5 * 1024 * 1024  # 5 MB upload limit

API_URL = os.environ['TCS_ION_API_URL']
USE_LOCAL_DATA = os.environ.get('USE_LOCAL_DATA', 'false').lower() == 'true'

# API timeout configuration (in seconds)
# The API typically takes 3+ minutes to respond, so set timeout accordingly
API_TIMEOUT = 360  # 360 seconds = 6 minutes (API can take 3-5 min; give headroom)

# Cache configuration
AUTO_REFRESH_INTERVAL_MINUTES = 30  # Auto-refresh cache every 5 minutes
CACHE_FILE = 'data/api_cache.json'  # Persistent cache file
_data_cache = None
_cache_timestamp = None
_cache_lock = threading.Lock()
_refresh_in_progress = False

# Secure password storage (hashed, loaded from env)
USERS = {
    'superadmin': generate_password_hash(os.environ['SUPERADMIN_PASSWORD']),
    'admin': generate_password_hash(os.environ['ADMIN_PASSWORD']),
    'user': generate_password_hash(os.environ['USER_PASSWORD']),
}
USER_ROLES = {'superadmin': 'superadmin', 'admin': 'admin', 'user': 'user'}
# Access model:
#   superadmin -> Dashboard + FY Analytics + Settings
#   admin      -> FY Analytics + Settings (no Dashboard)
#   user       -> FY Analytics only

# Login rate limiting: max 10 attempts per 60 seconds per IP
_login_attempts: dict = {}
_login_attempts_lock = threading.Lock()
_LOGIN_MAX = 10
_LOGIN_WINDOW = 60  # seconds


def _is_rate_limited(ip: str) -> bool:
    now = time.time()
    with _login_attempts_lock:
        attempts = _login_attempts.get(ip, [])
        attempts = [t for t in attempts if now - t < _LOGIN_WINDOW]
        _login_attempts[ip] = attempts
        if len(attempts) >= _LOGIN_MAX:
            return True
        attempts.append(now)
        _login_attempts[ip] = attempts
        return False


# Per-username soft lockout: N consecutive failures → temporary cooldown
_failed_logins: dict = {}  # username -> {'count': int, 'locked_until': float}
_failed_logins_lock = threading.Lock()
_LOCKOUT_THRESHOLD = 5
_LOCKOUT_SECONDS = 15 * 60  # 15-minute cooldown


def _is_locked_out(username: str) -> bool:
    if not username:
        return False
    with _failed_logins_lock:
        rec = _failed_logins.get(username)
        if not rec:
            return False
        if rec.get('locked_until', 0) > time.time():
            return True
        if rec.get('locked_until', 0):  # cooldown expired → clear
            _failed_logins.pop(username, None)
        return False


def _record_failed_login(username: str):
    if not username:
        return
    with _failed_logins_lock:
        rec = _failed_logins.setdefault(username, {'count': 0, 'locked_until': 0})
        rec['count'] += 1
        if rec['count'] >= _LOCKOUT_THRESHOLD:
            rec['locked_until'] = time.time() + _LOCKOUT_SECONDS
            rec['count'] = 0


def _reset_failed_login(username: str):
    with _failed_logins_lock:
        _failed_logins.pop(username, None)


def _is_valid_excel(file_bytes: bytes) -> bool:
    """Verify file content is a real Excel workbook by magic bytes,
    not just the filename extension. .xlsx = ZIP (PK..); .xls = OLE2."""
    return (
        file_bytes[:4] == b'PK\x03\x04'
        or file_bytes[:8] == b'\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1'
    )


# CSRF helpers
def _get_csrf_token() -> str:
    if 'csrf_token' not in session:
        session['csrf_token'] = secrets.token_hex(32)
    return session['csrf_token']


@app.context_processor
def inject_csrf_token():
    return {'csrf_token': _get_csrf_token()}


def csrf_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.headers.get('X-CSRF-Token') or request.form.get('csrf_token')
        if not token or token != session.get('csrf_token'):
            return jsonify({'error': 'CSRF validation failed'}), 403
        return f(*args, **kwargs)
    return decorated


def json_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not request.is_json:
            return jsonify({'error': 'Content-Type must be application/json'}), 415
        return f(*args, **kwargs)
    return decorated


# Security headers
@app.after_request
def add_security_headers(response):
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    response.headers['Permissions-Policy'] = 'geolocation=(), microphone=(), camera=()'
    response.headers['Content-Security-Policy'] = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline'; "
        "style-src 'self' 'unsafe-inline'; "
        "img-src 'self' data:; "
        "font-src 'self'; "
        "connect-src 'self';"
    )
    return response


# Session idle timeout (30 minutes of inactivity clears session)
@app.before_request
def check_session_idle():
    if 'logged_in' in session:
        last_active = session.get('_last_active')
        if last_active:
            idle_seconds = (datetime.now() - datetime.fromisoformat(last_active)).total_seconds()
            if idle_seconds > 1800:
                session.clear()
                return redirect(url_for('login'))
        session['_last_active'] = datetime.now().isoformat()


# Login required decorator
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function


# Admin required decorator (admin and superadmin)
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session:
            return redirect(url_for('login'))
        if session.get('role') not in ('admin', 'superadmin'):
            return jsonify({'error': 'Admin access required'}), 403
        return f(*args, **kwargs)
    return decorated_function


# Superadmin required decorator (dashboard APIs)
def superadmin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session:
            return redirect(url_for('login'))
        if session.get('role') != 'superadmin':
            return jsonify({'error': 'Superadmin access required'}), 403
        return f(*args, **kwargs)
    return decorated_function

# Helper functions for data persistence.
# Users and assignments now live in SQLite (see db.py); these two helpers keep
# their original signatures so every existing call site works unchanged, but they
# dispatch to the database by filename. Any other filename still falls back to
# plain JSON on disk (e.g. bundled/legacy data).
def load_json_file(filename, default=None):
    """Load a data store (SQLite-backed for users/assignments; JSON otherwise)."""
    try:
        if filename == USERS_FILE:
            return db.read_users()
        if filename == ASSIGNMENTS_FILE:
            return db.read_assignments()
        if os.path.exists(filename):
            with open(filename, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception as e:
        print(f"Error loading {filename}: {e}")
    return default if default is not None else {}

def save_json_file(filename, data):
    """Persist a data store (SQLite-backed for users/assignments; JSON otherwise)."""
    try:
        if filename == USERS_FILE:
            return db.write_users(data)
        if filename == ASSIGNMENTS_FILE:
            return db.write_assignments(data)
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        return True
    except Exception as e:
        print(f"Error saving {filename}: {e}")
        return False

# Data stores. The filenames double as the dispatch keys above and as the source
# for the one-time JSON -> SQLite bootstrap in db.init_db().
USERS_FILE = 'data/users.json'
ASSIGNMENTS_FILE = 'data/course_assignments.json'

# Ensure data directory exists, then open/initialise the SQLite database. On a
# brand-new app.db this also imports any existing JSON data exactly once.
os.makedirs('data', exist_ok=True)
db.init_db()


def _next_assignment_id(assignments):
    """Next unique assignment id = (highest existing id) + 1.

    Deliberately NOT len(assignments) + 1: assignments get deleted, which leaves
    gaps in the id sequence, so len()+1 can equal an id that is still live and
    silently mint a *duplicate* id. Lookups do next((a for a ... if id == x)) and
    return the first match, so a duplicate makes one assignment shadow another
    (e.g. a brand-new assignment showing a stale one's completion %)."""
    return max((a.get('id') or 0 for a in assignments), default=0) + 1


def _normalize_dt(value):
    """Normalize a date/datetime string (including the 'YYYY-MM-DDThh:mm' that HTML
    <input type="datetime-local"> emits) to canonical 'YYYY-MM-DD HH:MM:SS'.
    Returns '' if empty or unparseable."""
    dt = aa.parse_created(str(value or '').strip())
    return dt.strftime('%Y-%m-%d %H:%M:%S') if dt else ''

# Default deadline when an assignment is created without one: 15 days from creation.
DEFAULT_DEADLINE_DAYS = 15


def _default_deadline(created_dt=None):
    """Deadline to use when the admin leaves the field blank: DEFAULT_DEADLINE_DAYS
    after the assignment is created. Returned in the 'YYYY-MM-DD' shape the deadline
    field carries everywhere else (emails, reminders, the UI)."""
    base = created_dt or datetime.now()
    return (base + timedelta(days=DEFAULT_DEADLINE_DAYS)).strftime('%Y-%m-%d')


def _user_profiles():
    """email(lowercased) -> {name, department, location, job_role, tracked} for the
    analytics engine. The uploaded roster is the authentic user list; anyone the API
    reports who was never uploaded comes back tracked=False (untracked)."""
    try:
        return aa.build_profile_index(load_json_file(USERS_FILE, {'users': []}).get('users', []))
    except Exception as e:
        logger.error(f"Error building user profile index: {e}")
        return {}


def sync_users_from_api():
    """Add users the API reports but the local registry has never seen.

    These arrive with source='api' and tracked=0 — "untracked". They are NOT part of
    the authentic roster (which is what the admin uploads/adds), so they carry no
    department / location / role and the dashboards can exclude them with one switch.
    An existing user is never touched here: a bulk upload that already augmented them
    with a profile keeps it, and keeps tracked=1."""
    try:
        # Load API data
        data = load_data()

        # Extract unique user emails from API
        api_users = {}
        for record in data:
            email = record.get('User_Mail_ID', '').strip()
            name = record.get('Participant_Name', '').strip()
            if email and email.lower() not in api_users:
                api_users[email.lower()] = (email, name)

        # Load existing local users
        users_data = load_json_file(USERS_FILE, {'users': []})
        existing_users = users_data.get('users', [])

        # Case-insensitive lookup: the API and an uploaded roster can spell the same
        # address with different casing, and they must not become two user rows.
        existing_emails = {(u.get('email') or '').strip().lower() for u in existing_users}

        # Add new users from API that don't exist locally
        new_users_added = 0
        for key, (email, name) in api_users.items():
            if key not in existing_emails:
                existing_users.append({
                    'email': email,
                    'name': name,
                    'source': 'api',
                    'added_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'department': '',
                    'location': '',
                    'job_role': '',
                    'tracked': False,
                })
                new_users_added += 1

        # Save updated users list
        if new_users_added > 0:
            users_data['users'] = existing_users
            save_json_file(USERS_FILE, users_data)
            logger.info(f"Synced {new_users_added} new (untracked) users from API")

        return new_users_added
    except Exception as e:
        logger.error(f"Error syncing users from API: {e}")
        return 0

@app.route('/favicon.ico')
def favicon():
    return '', 204

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        ip = request.remote_addr
        if _is_rate_limited(ip):
            logger.warning(f"Rate limit exceeded for login from {ip}")
            return render_template('login.html', error='Too many login attempts. Please wait a moment.'), 429

        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')

        if _is_locked_out(username):
            logger.warning(f"Login blocked: account '{username}' temporarily locked (from {ip})")
            return render_template('login.html', error='Account temporarily locked due to repeated failed attempts. Try again later.'), 429

        if username in USERS and check_password_hash(USERS[username], password):
            _reset_failed_login(username)
            session.clear()
            session.permanent = True
            session['logged_in'] = True
            session['username'] = username
            session['role'] = USER_ROLES.get(username, 'user')
            session['_last_active'] = datetime.now().isoformat()
            # Superadmin lands on the Dashboard; admin/user land on FY Analytics
            if session['role'] == 'superadmin':
                return redirect(url_for('index'))
            return redirect(url_for('assignments_dashboard'))
        else:
            _record_failed_login(username)
            logger.warning(f"Failed login attempt for user '{username}' from {ip}")
            return render_template('login.html', error='Invalid username or password')

    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/settings')
@login_required
def settings():
    # Settings page is for admin and superadmin; others go to FY Analytics
    if session.get('role') not in ('admin', 'superadmin'):
        return redirect(url_for('assignments_dashboard'))
    return render_template('settings.html', username=session.get('username'),
                           role=session.get('role'))

# API endpoints for settings
@app.route('/api/settings/users', methods=['GET'])
@admin_required
def get_users():
    """Get all users (synced with API data)"""
    # Sync users from API first
    sync_users_from_api()
    
    # Return all users
    users_data = load_json_file(USERS_FILE, {'users': []})
    return jsonify(users_data.get('users', []))

@app.route('/api/settings/users', methods=['POST'])
@admin_required
@csrf_required
@json_required
def add_user():
    """Add a new user to the tracked roster (optionally with their profile)."""
    data = request.get_json()
    email = (data.get('email') or '').strip()

    if not email:
        return jsonify({'error': 'Email is required'}), 400

    users_data = load_json_file(USERS_FILE, {'users': []})
    users = users_data.get('users', [])

    # Check if user already exists (case-insensitive)
    email_lower = email.lower()
    existing = next((u for u in users if (u.get('email') or '').lower() == email_lower), None)
    if existing:
        # An API-synced (untracked) row is not a real duplicate — it is the same
        # person waiting for a profile. Promote it to the roster instead of
        # rejecting the admin's entry.
        if existing.get('tracked'):
            return jsonify({'error': f'User with email {email} already exists'}), 409
        db.upsert_user_profiles([{
            'email': existing['email'],
            'name': (data.get('name') or '').strip(),
            'department': (data.get('department') or '').strip(),
            'location': (data.get('location') or '').strip(),
            'job_role': (data.get('job_role') or '').strip(),
        }])
        updated = next((u for u in db.read_users().get('users', [])
                        if (u.get('email') or '').lower() == email_lower), None)
        return jsonify({'message': 'Existing API user promoted to tracked roster',
                        'user': updated, 'promoted': True})

    # Add new user
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    users.append({
        'email': email,
        'name': (data.get('name') or '').strip(),
        'source': 'manual',
        'added_date': now,
        'department': (data.get('department') or '').strip(),
        'location': (data.get('location') or '').strip(),
        'job_role': (data.get('job_role') or '').strip(),
        'tracked': True,
        'updated_date': now,
    })

    users_data['users'] = users
    if save_json_file(USERS_FILE, users_data):
        return jsonify({'message': 'User added successfully', 'user': users[-1]})
    else:
        return jsonify({'error': 'Failed to save user'}), 500

# Accepted spreadsheet headings for each roster field, lowercased. The upload is
# matched on these rather than on column position, so admins can reorder columns or
# omit the optional ones entirely.
_ROSTER_HEADERS = {
    'email': ('email', 'e-mail', 'email address', 'email id', 'mail', 'user_mail_id',
              'user mail id'),
    'name': ('name', 'full name', 'participant name', 'employee name', 'user name'),
    'department': ('department', 'dept', 'function', 'business unit'),
    'location': ('location', 'site', 'office', 'branch', 'plant', 'city'),
    'job_role': ('role', 'job role', 'designation', 'position', 'job title', 'title',
                 'grade'),
}
# Fallback when the sheet has no recognisable header row (the legacy format was a
# bare column of emails, so column A must stay the email column).
_ROSTER_POSITIONAL = ('email', 'name', 'department', 'location', 'job_role')


def _roster_column_map(first_row):
    """Map column index -> roster field from a sheet's first row, or None if that row
    is not a header (i.e. it already holds data)."""
    mapping = {}
    for idx, cell in enumerate(first_row or ()):
        text = str(cell or '').strip().lower()
        if not text:
            continue
        for field, aliases in _ROSTER_HEADERS.items():
            if text in aliases and field not in mapping.values():
                mapping[idx] = field
                break
    # A header row must at least name the email column; otherwise treat the sheet as
    # headerless data in the legacy positional layout.
    return mapping if 'email' in mapping.values() else None


def _parse_roster_sheet(sheet):
    """Parse an uploaded roster into ({email, name, department, location, job_role}, ...)
    plus the list of unparseable email values. Understands both a headed sheet (in any
    column order) and the legacy headerless 'emails in column A' sheet."""
    email_pattern = re.compile(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$')
    rows = list(sheet.iter_rows(min_row=1, values_only=True))
    if not rows:
        return [], []

    colmap = _roster_column_map(rows[0])
    if colmap is not None:
        data_rows = rows[1:]
    else:
        colmap = {i: f for i, f in enumerate(_ROSTER_POSITIONAL)}
        data_rows = rows

    parsed, invalid = [], []
    for row in data_rows:
        record = {f: '' for f in _ROSTER_POSITIONAL}
        for idx, field in colmap.items():
            if idx < len(row):
                record[field] = str(row[idx] or '').strip()
        email = record['email']
        if not email:
            continue  # blank/spacer row
        if not email_pattern.match(email):
            invalid.append(email)
            continue
        parsed.append(record)
    return parsed, invalid


@app.route('/api/settings/users/bulk-upload', methods=['POST'])
@admin_required
@csrf_required
def bulk_upload_users():
    """Import the authentic user roster from Excel — email plus optional name,
    department, location and role.

    The upload is the source of truth for *who* is being tracked and for their
    profile. An email already in the registry (including one the API sync created) is
    augmented in place with whatever the sheet supplies and promoted to tracked;
    unknown emails are inserted. Users the sheet does not mention are left exactly as
    they are — an API-only user stays untracked, with no profile."""
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400

    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Invalid file format. Please upload an Excel file (.xlsx or .xls)'}), 400

    file_bytes = file.read()
    if not _is_valid_excel(file_bytes):
        return jsonify({'error': 'Invalid file content. File is not a valid Excel workbook.'}), 400

    try:
        workbook = openpyxl.load_workbook(BytesIO(file_bytes))
        sheet = workbook.active
        parsed, invalid_emails = _parse_roster_sheet(sheet)

        if not parsed:
            return jsonify({
                'error': 'No valid email addresses found in file',
                'invalid_emails': invalid_emails
            }), 400

        # Last row wins if the same person appears twice in one sheet.
        deduped = {}
        for r in parsed:
            deduped[r['email'].lower()] = r

        existing = {(u.get('email') or '').lower()
                    for u in load_json_file(USERS_FILE, {'users': []}).get('users', [])}
        added_users = [r['email'] for k, r in deduped.items() if k not in existing]
        augmented_users = [r['email'] for k, r in deduped.items() if k in existing]

        result = db.upsert_user_profiles(list(deduped.values()))
        profiled = sum(
            1 for r in deduped.values()
            if r['department'] or r['location'] or r['job_role'])

        logger.info(
            f"Roster upload by {session.get('username')}: "
            f"{result['added']} added, {result['updated']} augmented, "
            f"{len(invalid_emails)} invalid")

        return jsonify({
            'message': f'Successfully processed {len(deduped)} user(s)',
            'added_count': result['added'],
            'added_users': added_users,
            'updated_count': result['updated'],
            'updated_users': augmented_users,
            'profiled_count': profiled,
            # Kept for the existing UI copy: an email already on file is no longer
            # "skipped as a duplicate", it is augmented — so report it as such.
            'duplicate_count': result['updated'],
            'duplicate_users': augmented_users,
            'invalid_count': len(invalid_emails),
            'invalid_emails': invalid_emails
        })

    except Exception as e:
        logger.error(f"Error processing Excel file (users): {e}")
        return jsonify({'error': 'Failed to process file'}), 500

@app.route('/api/settings/users/<email>', methods=['DELETE'])
@admin_required
@csrf_required
def delete_user(email):
    """Delete a user"""
    users_data = load_json_file(USERS_FILE, {'users': []})
    users = users_data.get('users', [])
    
    # Remove user
    users = [u for u in users if u['email'] != email]
    users_data['users'] = users
    
    if save_json_file(USERS_FILE, users_data):
        return jsonify({'message': 'User deleted successfully'})
    else:
        return jsonify({'error': 'Failed to delete user'}), 500

@app.route('/api/settings/courses', methods=['GET'])
@admin_required
def get_available_courses():
    """Get list of available courses from the data"""
    try:
        data = load_data()
        # Get unique course names
        courses = list(set(record.get('Activity_Name', record.get('Course_Name', '')) for record in data))
        courses = [c for c in courses if c]  # Remove empty strings
        courses.sort()
        return jsonify({'courses': courses})
    except Exception as e:
        logger.error(f"Error loading courses: {e}")
        return jsonify({'error': 'Failed to load courses'}), 500

# ── Email Job Tracking ────────────────────────────────────────────────────────
# Each job: {total, sent, failed, status: 'running'|'completed'}
_email_jobs: dict = {}
_email_jobs_lock = threading.Lock()

_EMAIL_WORKERS = 8   # parallel SMTP connections per job
_JOB_TTL = 600       # seconds before a completed job is cleaned up


def _create_job(total: int) -> str:
    job_id = str(uuid.uuid4())
    with _email_jobs_lock:
        _email_jobs[job_id] = {'total': total, 'sent': 0, 'failed': 0, 'status': 'running'}
    return job_id


def _mark_job_done(job_id: str):
    with _email_jobs_lock:
        if job_id in _email_jobs:
            _email_jobs[job_id]['status'] = 'completed'
    def _cleanup():
        time.sleep(_JOB_TTL)
        with _email_jobs_lock:
            _email_jobs.pop(job_id, None)
    threading.Thread(target=_cleanup, daemon=True).start()


def _run_email_job(job_id, emails, send_fn):
    """Execute a list of (email, name, ...) send calls in parallel and track progress."""
    def _one(task):
        try:
            return send_fn(task)
        except Exception as e:
            logger.error(f"Email job {job_id}: error: {e}")
            return False

    with ThreadPoolExecutor(max_workers=_EMAIL_WORKERS) as pool:
        futures = {pool.submit(_one, task): task for task in emails}
        for future in as_completed(futures):
            ok = future.result()
            with _email_jobs_lock:
                if job_id in _email_jobs:
                    if ok:
                        _email_jobs[job_id]['sent'] += 1
                    else:
                        _email_jobs[job_id]['failed'] += 1
    _mark_job_done(job_id)


def _dispatch_assignment_emails(emails, course_name, deadline, email_to_name) -> str:
    """Start a background job to send assignment notification emails. Returns job_id."""
    tasks = [(email, email_to_name.get(email, email.split('@')[0])) for email in emails]
    job_id = _create_job(len(tasks))

    def _send(task):
        user_email, user_name = task
        return send_course_assignment_email(user_email, user_name, course_name, deadline)

    t = threading.Thread(target=_run_email_job, args=(job_id, tasks, _send), daemon=True)
    t.start()
    return job_id


def _dispatch_removal_emails(emails, course_name, email_to_name) -> str:
    """Start a background job to send removal notification emails. Returns job_id."""
    tasks = [(email, email_to_name.get(email, email.split('@')[0])) for email in emails]
    job_id = _create_job(len(tasks))

    def _send(task):
        user_email, user_name = task
        return send_course_removal_email(user_email, user_name, course_name)

    t = threading.Thread(target=_run_email_job, args=(job_id, tasks, _send), daemon=True)
    t.start()
    return job_id


@app.route('/api/email-job/<job_id>', methods=['GET'])
@admin_required
def get_email_job_status(job_id):
    """Return current progress of an email dispatch job."""
    with _email_jobs_lock:
        job = _email_jobs.get(job_id)
    if not job:
        return jsonify({'error': 'Job not found or already expired'}), 404
    return jsonify(job)


@app.route('/api/settings/assignments', methods=['GET'])
@admin_required
def get_assignments():
    """Get all course assignments"""
    assignments = load_json_file(ASSIGNMENTS_FILE, {'assignments': []})
    return jsonify(assignments.get('assignments', []))

@app.route('/api/settings/assignments', methods=['POST'])
@admin_required
@csrf_required
@json_required
def create_assignment():
    """Create a new course assignment"""
    data = request.get_json()
    course_name = data.get('course_name')
    user_emails = data.get('user_emails', [])[:500]
    deadline = (data.get('deadline') or '').strip()
    notify_email = data.get('notify_email', False)

    if not course_name or not user_emails:
        return jsonify({'error': 'Course name and users are required'}), 400

    # Validity window: completions only count within [effective_from, effective_to].
    # effective_from defaults to the creation time; effective_to is optional (open).
    created_dt = datetime.now()
    created = created_dt.strftime('%Y-%m-%d %H:%M:%S')
    effective_from = _normalize_dt(data.get('effective_from')) or created
    effective_to = _normalize_dt(data.get('effective_to'))
    if effective_to and effective_from > effective_to:
        return jsonify({'error': 'Effective-from must be on or before effective-to'}), 400

    # No deadline given → 15 days from now (DEFAULT_DEADLINE_DAYS).
    if not deadline:
        deadline = _default_deadline(created_dt)

    assignments_data = load_json_file(ASSIGNMENTS_FILE, {'assignments': []})
    assignments = assignments_data.get('assignments', [])

    # Create new assignment
    assignment = {
        'id': _next_assignment_id(assignments),
        'course_name': course_name,
        'user_emails': user_emails,
        'deadline': deadline,
        'effective_from': effective_from,
        'effective_to': effective_to,
        'created_date': created,
        'created_by': session.get('username')
    }
    assignment['title'] = aa.assignment_title(assignment)

    assignments.append(assignment)
    assignments_data['assignments'] = assignments

    if not save_json_file(ASSIGNMENTS_FILE, assignments_data):
        return jsonify({'error': 'Failed to save assignment'}), 500

    # Start watching for completions. Anyone already complete right now (which a
    # backdated effective_from deliberately credits) is recorded as pre-existing so
    # they are not congratulated for something they finished before this existed.
    baseline_assignment_notifications(assignment)

    # Dispatch email notifications in background if requested
    email_job_id = None
    if notify_email:
        users_data = load_json_file(USERS_FILE, {'users': []})
        email_to_name = {u['email']: u['name'] for u in users_data.get('users', [])}
        email_job_id = _dispatch_assignment_emails(user_emails, course_name, deadline, email_to_name)

    return jsonify({
        'message': 'Assignment created successfully',
        'assignment': assignment,
        'deadline': deadline,
        'deadline_defaulted': not (data.get('deadline') or '').strip(),
        'notify_email': notify_email,
        'email_job_id': email_job_id,
        'email_total': len(user_emails) if notify_email else 0
    }), 201

@app.route('/api/settings/assignments/<int:assignment_id>', methods=['GET'])
@admin_required
def get_assignment_details(assignment_id):
    """Get detailed assignment statistics with course data"""
    try:
        # Load assignment
        assignments_data = load_json_file(ASSIGNMENTS_FILE, {'assignments': []})
        assignment = next((a for a in assignments_data.get('assignments', []) if a.get('id') == assignment_id), None)
        
        if not assignment:
            return jsonify({'error': 'Assignment not found'}), 404
        
        # Load API data
        api_data = load_data()
        course_name = assignment['course_name']

        # Get all users from API for this course (Course View)
        course_records = [r for r in api_data if r.get('Activity_Name', r.get('Course_Name')) == course_name]
        
        # Course View: All users in API for this course
        course_users = {}
        for record in course_records:
            email = record.get('User_Mail_ID', '')
            if email and email not in course_users:
                course_users[email] = {
                    'email': email,
                    'name': record.get('Participant_Name', ''),
                    'completion_percentage': record.get('Course_Completion_Percentage', '0'),
                    'completion_status': record.get('Course_Completion_Status', ''),
                    'activity_status': record.get('Activity_Status', ''),
                    'completion_date': record.get('Course_Completion_Date_(YYYY-MM-DD)', '').strip()
                }
        
        # Assignment View: staleness-aware. A completion counts only if it is dated
        # on/after the assignment was created; completions that predate the assignment
        # are "stale" (the user must complete it again) and do NOT count as done.
        # Uses the same engine as the FY dashboard so both views agree.
        api_index = aa.build_api_index(api_data)
        prog = aa.compute_assignment_progress(assignment, api_index)
        status_display = {'completed': 'Completed', 'in_progress': 'Current',
                          'stale': 'Stale', 'not_started': 'Not Started'}
        assignment_users = []
        for u in prog['users']:
            assignment_users.append({
                'email': u['email'],
                'name': u['name'],
                # a stale completion contributes 0% toward this assignment
                'completion_percentage': '0' if u['stale'] else str(u['percentage']),
                'completion_status': status_display[u['status']],
                'activity_status': '',
                'completion_date': u['completion_date'],
                'untouched': u['untouched'],
                'stale': u['stale'],
            })

        # Course View stats (course-wide, raw API — no assignment-date staleness)
        course_completed = sum(1 for u in course_users.values() if u['completion_status'] == 'Completed')
        course_in_progress = sum(1 for u in course_users.values() if u['completion_status'] == 'Current')
        course_not_started = len(course_users) - course_completed - course_in_progress
        course_completion_rate = (course_completed / len(course_users) * 100) if len(course_users) > 0 else 0

        return jsonify({
            'assignment': assignment,
            'course_view': {
                'total_users': len(course_users),
                'completed': course_completed,
                'in_progress': course_in_progress,
                'not_started': course_not_started,
                'stale': 0,
                'completion_rate': round(course_completion_rate, 2),
                'users': list(course_users.values())
            },
            'assignment_view': {
                'total_users': prog['total'],
                'completed': prog['completed'],
                'in_progress': prog['in_progress'],
                'not_started': prog['not_started'],
                'stale': prog['stale'],
                'completion_rate': prog['completion_rate'],
                'users': assignment_users
            }
        })
    except Exception as e:
        logger.error(f"Error getting assignment details: {e}")
        return jsonify({'error': 'Failed to load assignment details'}), 500

@app.route('/api/settings/assignments/<int:assignment_id>/remind', methods=['POST'])
@admin_required
@csrf_required
def send_assignment_reminders(assignment_id):
    """Send reminder emails to incomplete users"""
    try:
        from datetime import datetime, timedelta
        from email_service import send_deadline_reminder_email
        
        # Load assignment
        assignments_data = load_json_file(ASSIGNMENTS_FILE, {'assignments': []})
        assignment = next((a for a in assignments_data.get('assignments', []) if a.get('id') == assignment_id), None)
        
        if not assignment:
            return jsonify({'error': 'Assignment not found'}), 404
        
        # Load API data to check completion status
        api_data = load_data()
        course_name = assignment['course_name']
        assigned_users = assignment['user_emails']
        deadline = assignment['deadline']
        
        # Calculate days remaining
        try:
            deadline_date = datetime.strptime(deadline, '%Y-%m-%d')
            today = datetime.now()
            days_remaining = (deadline_date - today).days
        except:
            days_remaining = 0
        
        # Only VALID completions (dated on/after the assignment) are skipped.
        # A stale completion means the user must redo the course, so they still
        # get a reminder — same staleness rule as the dashboards.
        _prog = aa.compute_assignment_progress(assignment, aa.build_api_index(api_data))
        completed_users = {u['email'] for u in _prog['users'] if u['status'] == 'completed'}
        
        # Get users data for names
        users_data = load_json_file(USERS_FILE, {'users': []})
        email_to_name = {user['email']: user['name'] for user in users_data.get('users', [])}
        
        # Send reminders to incomplete users only
        results = {'success': 0, 'failed': 0, 'skipped': 0, 'failed_emails': []}
        
        for user_email in assigned_users:
            if user_email in completed_users:
                results['skipped'] += 1
                continue
            
            user_name = email_to_name.get(user_email, user_email.split('@')[0])
            
            try:
                success = send_deadline_reminder_email(
                    user_email=user_email,
                    user_name=user_name,
                    course_name=course_name,
                    deadline=deadline,
                    days_remaining=days_remaining
                )
                
                if success:
                    results['success'] += 1
                    logger.info(f"Reminder sent to {user_email}")
                else:
                    results['failed'] += 1
                    results['failed_emails'].append(user_email)
                    logger.warning(f"Failed to send reminder to {user_email}")
            except Exception as e:
                results['failed'] += 1
                results['failed_emails'].append(user_email)
                logger.error(f"Error sending reminder to {user_email}: {e}")
        
        return jsonify({
            'message': 'Reminders sent',
            'reminders_sent': results['success'],
            'reminders_failed': results['failed'],
            'already_completed': results['skipped'],
            'failed_emails': results['failed_emails']
        })
    except Exception as e:
        logger.error(f"Error sending reminders: {e}")
        return jsonify({'error': 'Failed to send reminders'}), 500

@app.route('/api/settings/assignments/<int:assignment_id>', methods=['PUT'])
@admin_required
@csrf_required
@json_required
def update_assignment(assignment_id):
    """Update assignment (add/remove users)"""
    try:
        from email_service import send_course_assignment_email, send_course_removal_email
        
        data = request.get_json()
        new_user_emails = data.get('user_emails', [])[:500]
        notify_email = data.get('notify_email', False)

        # Load assignment
        assignments_data = load_json_file(ASSIGNMENTS_FILE, {'assignments': []})
        assignments = assignments_data.get('assignments', [])
        assignment = next((a for a in assignments if a.get('id') == assignment_id), None)

        if not assignment:
            return jsonify({'error': 'Assignment not found'}), 404

        old_user_emails = set(assignment['user_emails'])
        new_user_emails_set = set(new_user_emails)

        # Find added and removed users
        added_users = new_user_emails_set - old_user_emails
        removed_users = old_user_emails - new_user_emails_set

        # Update assignment
        assignment['user_emails'] = new_user_emails
        assignment['last_modified'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # Optionally edit the validity window (only if the field is present in the
        # payload). effective_from can't be blank — it falls back to created_date;
        # effective_to may be cleared to '' to mean "open-ended".
        if 'effective_from' in data:
            assignment['effective_from'] = (
                _normalize_dt(data.get('effective_from')) or assignment.get('created_date', ''))
        if 'effective_to' in data:
            assignment['effective_to'] = _normalize_dt(data.get('effective_to'))
        ef, et = assignment.get('effective_from', ''), assignment.get('effective_to', '')
        if ef and et and ef > et:
            return jsonify({'error': 'Effective-from must be on or before effective-to'}), 400

        # Save first so response is immediate
        if not save_json_file(ASSIGNMENTS_FILE, assignments_data):
            return jsonify({'error': 'Failed to save assignment'}), 500

        # Keep the completion-notification ledger in step with the roster change:
        #  • a user ADDED who has already completed the course inside the window is
        #    suppressed — they finished before they were assigned it, so there is no
        #    completion event to congratulate (same rule as the creation baseline);
        #  • a user REMOVED has their row dropped, so if they are re-added later and
        #    then complete, that genuinely new completion is still announced.
        sync_assignment_notifications(assignment, added_users, removed_users)

        # Dispatch email notifications in background if requested
        email_job_id = None
        email_total = 0
        if notify_email:
            users_data = load_json_file(USERS_FILE, {'users': []})
            email_to_name = {u['email']: u['name'] for u in users_data.get('users', [])}
            recipients = list(added_users) + list(removed_users)
            email_total = len(recipients)
            if added_users:
                email_job_id = _dispatch_assignment_emails(
                    list(added_users), assignment['course_name'], assignment['deadline'], email_to_name)
            if removed_users:
                _dispatch_removal_emails(list(removed_users), assignment['course_name'], email_to_name)

        return jsonify({
            'message': 'Assignment updated successfully',
            'assignment': assignment,
            'users_added': len(added_users),
            'users_removed': len(removed_users),
            'notify_email': notify_email,
            'email_job_id': email_job_id,
            'email_total': email_total
        })
    except Exception as e:
        logger.error(f"Error updating assignment: {e}")
        return jsonify({'error': 'Failed to update assignment'}), 500

@app.route('/api/settings/assignments/<int:assignment_id>', methods=['DELETE'])
@admin_required
@csrf_required
def delete_assignment(assignment_id):
    """Delete a course assignment"""
    assignments_data = load_json_file(ASSIGNMENTS_FILE, {'assignments': []})
    assignments = assignments_data.get('assignments', [])

    # Remove assignment
    assignments = [a for a in assignments if a.get('id') != assignment_id]
    assignments_data['assignments'] = assignments

    if save_json_file(ASSIGNMENTS_FILE, assignments_data):
        # Drop the completion-notification rows too. Ids are minted as max+1, so a
        # deleted id can never be reused — but leaving the rows behind would keep a
        # dead assignment's "already congratulated" state around forever.
        db.delete_notifications_for_assignment(assignment_id)
        return jsonify({'message': 'Assignment deleted successfully'})
    else:
        return jsonify({'error': 'Failed to delete assignment'}), 500

@app.route('/api/settings/assignments/bulk-upload', methods=['POST'])
@superadmin_required
@csrf_required
def bulk_upload_assignment():
    """Bulk upload users for course assignment from Excel file"""
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    
    course_name = request.form.get('course_name')
    deadline = (request.form.get('deadline') or '').strip()
    notify_email = request.form.get('notify_email', 'false').lower() == 'true'

    if not course_name:
        return jsonify({'error': 'Course name is required'}), 400

    # Validity window (optional): effective_from defaults to creation time below.
    created_dt = datetime.now()
    created = created_dt.strftime('%Y-%m-%d %H:%M:%S')
    effective_from = _normalize_dt(request.form.get('effective_from')) or created
    effective_to = _normalize_dt(request.form.get('effective_to'))
    if effective_to and effective_from > effective_to:
        return jsonify({'error': 'Effective-from must be on or before effective-to'}), 400

    # No deadline given → 15 days from now (DEFAULT_DEADLINE_DAYS).
    if not deadline:
        deadline = _default_deadline(created_dt)
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Invalid file format. Please upload an Excel file (.xlsx or .xls)'}), 400

    file_bytes = file.read()
    if not _is_valid_excel(file_bytes):
        return jsonify({'error': 'Invalid file content. File is not a valid Excel workbook.'}), 400

    try:
        # Read Excel file
        workbook = openpyxl.load_workbook(BytesIO(file_bytes))
        sheet = workbook.active

        # Email validation regex
        email_pattern = re.compile(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$')

        valid_emails = []
        invalid_emails = []

        # Parse rows for emails (check first column)
        for row in sheet.iter_rows(min_row=1, values_only=True):
            if row[0]:  # If first cell has value
                email = str(row[0]).strip()
                if email_pattern.match(email):
                    valid_emails.append(email)
                else:
                    invalid_emails.append(email)

        if not valid_emails:
            return jsonify({
                'error': 'No valid email addresses found in file',
                'invalid_emails': invalid_emails
            }), 400
        
        # Load existing users
        users_data = load_json_file(USERS_FILE, {'users': []})
        users = users_data.get('users', [])
        existing_emails = {u['email'].lower() for u in users}
        
        # Add new users if they don't exist. An email the admin explicitly assigns is
        # part of the roster, so it lands tracked (with an empty profile until a
        # roster upload fills in department / location / role).
        new_users_added = []
        for email in valid_emails:
            email_lower = email.lower()
            if email_lower not in existing_emails:
                now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                new_user = {
                    'email': email,
                    'name': '',
                    'source': 'manual',
                    'added_date': now,
                    'department': '',
                    'location': '',
                    'job_role': '',
                    'tracked': True,
                    'updated_date': now,
                }
                users.append(new_user)
                new_users_added.append(email)
                existing_emails.add(email_lower)
        
        # Save updated users if any new users were added
        if new_users_added:
            users_data['users'] = users
            save_json_file(USERS_FILE, users_data)
        
        # Create assignment with all valid emails
        assignments_data = load_json_file(ASSIGNMENTS_FILE, {'assignments': []})
        assignments = assignments_data.get('assignments', [])
        
        assignment = {
            'id': _next_assignment_id(assignments),
            'course_name': course_name,
            'user_emails': valid_emails,
            'deadline': deadline,
            'effective_from': effective_from,
            'effective_to': effective_to,
            'created_date': created,
            'created_by': session.get('username')
        }
        assignment['title'] = aa.assignment_title(assignment)

        assignments.append(assignment)
        assignments_data['assignments'] = assignments
        
        if not save_json_file(ASSIGNMENTS_FILE, assignments_data):
            return jsonify({'error': 'Failed to save assignment'}), 500

        # Watch for completions from here on (see create_assignment).
        baseline_assignment_notifications(assignment)

        # Dispatch email notifications in background if requested
        email_job_id = None
        if notify_email:
            email_to_name = {u['email']: u['name'] for u in users}
            email_job_id = _dispatch_assignment_emails(valid_emails, course_name, deadline, email_to_name)

        return jsonify({
            'message': 'Assignment created successfully',
            'assignment': assignment,
            'users_assigned': len(valid_emails),
            'new_users_added': len(new_users_added),
            'new_users_list': new_users_added,
            'invalid_count': len(invalid_emails),
            'invalid_emails': invalid_emails,
            'notify_email': notify_email,
            'email_job_id': email_job_id,
            'email_total': len(valid_emails) if notify_email else 0
        }), 201
            
    except Exception as e:
        logger.error(f"Error processing Excel file (assignments): {e}")
        return jsonify({'error': 'Failed to process file'}), 500

@app.route('/')
@login_required
def index():
    # Dashboard is superadmin-only; everyone else lands on FY Analytics
    if session.get('role') != 'superadmin':
        return redirect(url_for('assignments_dashboard'))
    return render_template('index.html', username=session.get('username'),
                           role=session.get('role'))

# Available to all logged-in users: the FY Analytics page shows cache
# freshness and offers the refresh button to every role.
@app.route('/api/status')
@login_required
def get_status():
    """Get system status including cache information"""
    cache_info = {
        'cached': _data_cache is not None,
        'cache_timestamp': None,
        'cache_age_seconds': None,
        'cache_age_readable': None,
        'refresh_in_progress': _refresh_in_progress
    }
    
    if _cache_timestamp is not None:
        now = datetime.now()
        cache_age = now - _cache_timestamp
        cache_age_seconds = int(cache_age.total_seconds())
        
        # Calculate readable age
        days = cache_age.days
        hours = cache_age.seconds // 3600
        minutes = (cache_age.seconds % 3600) // 60
        
        age_parts = []
        if days > 0:
            age_parts.append(f"{days} day{'s' if days != 1 else ''}")
        if hours > 0:
            age_parts.append(f"{hours} hour{'s' if hours != 1 else ''}")
        if minutes > 0 or len(age_parts) == 0:
            age_parts.append(f"{minutes} minute{'s' if minutes != 1 else ''}")
        
        cache_info['cache_timestamp'] = _cache_timestamp.strftime('%Y-%m-%d %H:%M:%S')
        cache_info['cache_age_seconds'] = cache_age_seconds
        cache_info['cache_age_readable'] = ' '.join(age_parts)
        cache_info['cached_at_formatted'] = _cache_timestamp.strftime('%d %B %Y at %I:%M %p')
    
    return jsonify({
        'using_local_data': USE_LOCAL_DATA,
        'data_source': 'Local JSON File' if USE_LOCAL_DATA else 'TCS iON API',
        'api_timeout_seconds': API_TIMEOUT,
        'auto_refresh_interval_minutes': AUTO_REFRESH_INTERVAL_MINUTES,
        'cache': cache_info
    })

@app.route('/api/refresh-cache', methods=['POST'])
@login_required
@csrf_required
def refresh_cache_endpoint():
    """Trigger manual cache refresh (non-blocking)"""
    global _refresh_in_progress
    
    if USE_LOCAL_DATA:
        return jsonify({'error': 'Cache refresh not available in local data mode'}), 400
    
    with _cache_lock:
        if _refresh_in_progress:
            return jsonify({
                'message': 'Cache refresh already in progress',
                'refresh_in_progress': True
            }), 202
    
    # Start refresh in background thread
    threading.Thread(target=refresh_cache_background, daemon=True).start()
    
    logger.info(f"Manual cache refresh triggered by {session.get('username')}")
    
    return jsonify({
        'message': 'Cache refresh started in background',
        'refresh_in_progress': True
    }), 202

def save_cache_to_file(data, timestamp):
    """Persist the API cache to SQLite (see db.py). Name kept for call sites."""
    try:
        if db.write_cache(data, timestamp):
            logger.info("Cache saved to SQLite")
            return True
        return False
    except Exception as e:
        logger.error(f"Error saving cache: {e}")
        return False

def load_cache_from_file():
    """Load the API cache from SQLite (see db.py). Name kept for call sites."""
    try:
        data, timestamp = db.read_cache()
        if data and timestamp:
            logger.info(f"Cache loaded from SQLite (cached at: {timestamp.strftime('%Y-%m-%d %H:%M:%S')})")
            return data, timestamp
    except Exception as e:
        logger.error(f"Error loading cache: {e}")
    return None, None

def record_completion_history(data, timestamp):
    """Fold an API snapshot's completions into the persistent ledger (db.py).
    Best-effort: never let a ledger hiccup break a cache refresh."""
    try:
        new = db.record_completions(aa.extract_completions(data), timestamp)
        if new:
            logger.info(f"Completion ledger: recorded {new} new completion date(s)")
        return new
    except Exception as e:
        logger.error(f"Error recording completion history: {e}")
        return 0


def _history_index():
    """(course, email) -> [completion datetimes] from the persistent ledger, for
    the assignment analytics engine. Empty on any error so the dashboard degrades
    to live-snapshot-only behaviour rather than failing."""
    try:
        return aa.build_history_index(db.read_completion_history())
    except Exception as e:
        logger.error(f"Error building completion-history index: {e}")
        return {}


# ── Completion notifications ────────────────────────────────────────────────
# Every API sync asks: has anyone completed a course that an assignment requires of
# them, and have we told them yet? The completion_notifications ledger answers the
# second half — one row per (assignment, user), so a congratulations email is sent
# exactly once no matter how many times the same completion is re-observed.
#
# Ordering matters. An assignment is "baselined" the moment it starts being watched
# (at creation, or on the first sync for assignments that predate this feature):
# whoever is already Completed then is recorded as pre_existing and never emailed.
# Only a completion that appears *after* the baseline is an event worth announcing.
NOTIFY_ON_COMPLETION = os.environ.get('NOTIFY_ON_COMPLETION', 'true').lower() == 'true'
_notify_lock = threading.Lock()


def _completion_candidates(assignment, api_index, history_index, email_to_name):
    """Everyone who currently counts as Completed for this assignment (i.e. finished
    inside its validity window), as notification-ledger candidate dicts."""
    prog = aa.compute_assignment_progress(assignment, api_index, history_index)
    out = []
    for u in prog['users']:
        if u['status'] != 'completed':
            continue
        email = u['email']
        out.append({
            'assignment_id': assignment.get('id'),
            'email': email,
            'course': prog['course_name'],
            'completion_date': u['completion_date'],
            'deadline': assignment.get('deadline') or '',
            'name': (u['name'] or email_to_name.get(email.lower()) or email.split('@')[0]),
        })
    return out


def _email_to_name():
    return {(u.get('email') or '').lower(): (u.get('name') or '')
            for u in load_json_file(USERS_FILE, {'users': []}).get('users', [])}


def _warm_data_or_none():
    """The cached API snapshot, or None if the cache is not warm yet.

    Unlike load_data(), this NEVER falls back to a synchronous API fetch — it is called
    on the assignment-creation request path, where a 3+ minute blocking fetch (which is
    what load_data() does on a cold start with no cache) would hang the admin's browser."""
    if USE_LOCAL_DATA:
        return load_data()
    with _cache_lock:
        return _data_cache


def baseline_assignment_notifications(assignment, api_data=None):
    """Start watching one assignment, suppressing congratulations for users who have
    already completed it. Called when an assignment is created — so a backdated
    validity window (which instantly credits old completions) does not fire a burst
    of emails at people who finished months ago."""
    if not NOTIFY_ON_COMPLETION:
        return 0
    try:
        data = api_data if api_data is not None else _warm_data_or_none()
        if data is None:
            # Cold start, cache not populated yet. Leave the assignment un-baselined:
            # the next sync's dispatcher baselines it against fresh data, which
            # suppresses whoever is already complete just the same.
            logger.info(
                f"Assignment {assignment.get('id')}: cache not warm, deferring "
                f"notification baseline to the next sync")
            return 0
        candidates = _completion_candidates(
            assignment, aa.build_api_index(data), _history_index(), _email_to_name())
        seeded = db.set_notification_baseline(assignment.get('id'), candidates)
        if seeded:
            logger.info(
                f"Assignment {assignment.get('id')}: {seeded} user(s) already complete "
                f"at creation — recorded, not emailed")
        return seeded
    except Exception as e:
        logger.error(f"Error baselining assignment notifications: {e}")
        return 0


def sync_assignment_notifications(assignment, added_users, removed_users):
    """Reconcile the notification ledger with an edited assignment roster (see the
    call site in update_assignment). Best-effort — never fails the update."""
    if not NOTIFY_ON_COMPLETION:
        return
    try:
        aid = assignment.get('id')
        for email in (removed_users or ()):
            db.settle_notification(aid, email, sent=False)   # deletes the row
        if not added_users:
            return
        data = _warm_data_or_none()
        if data is None:
            return   # cache cold; the next sync baselines/claims against fresh data
        added = set(added_users)
        already_done = [
            c for c in _completion_candidates(
                assignment, aa.build_api_index(data), _history_index(), _email_to_name())
            if c['email'] in added]
        n = db.suppress_completion_notifications(already_done)
        if n:
            logger.info(f"Assignment {aid}: {n} newly-added user(s) already complete "
                        f"— recorded, not emailed")
    except Exception as e:
        logger.error(f"Error syncing assignment notifications: {e}")


def dispatch_completion_notifications(api_data=None):
    """Congratulate every user who has newly completed an assigned course.

    Runs after each cache refresh. Non-blocking for callers that hold no lock: a pass
    already under way makes this a no-op rather than queueing a second one."""
    if not NOTIFY_ON_COMPLETION:
        return {'sent': 0, 'failed': 0, 'skipped': 0}
    if not _notify_lock.acquire(blocking=False):
        logger.info("Completion notification pass already running, skipping")
        return {'sent': 0, 'failed': 0, 'skipped': 0}

    try:
        data = api_data if api_data is not None else load_data()
        assignments = load_json_file(ASSIGNMENTS_FILE, {'assignments': []}).get('assignments', [])
        if not assignments:
            return {'sent': 0, 'failed': 0, 'skipped': 0}

        api_index = aa.build_api_index(data)
        history_index = _history_index()
        email_to_name = _email_to_name()
        baselined = db.baselined_assignment_ids()

        candidates = []
        for a in assignments:
            aid = a.get('id')
            done = _completion_candidates(a, api_index, history_index, email_to_name)
            if aid not in baselined:
                # First time this assignment is watched (it predates the feature, or
                # was created while notifications were switched off). Everyone already
                # finished is recorded silently; only later completions get a mail.
                db.set_notification_baseline(aid, done)
                continue
            candidates.extend(done)

        claimed = db.claim_completion_notifications(candidates)
        if not claimed:
            return {'sent': 0, 'failed': 0, 'skipped': 0}

        logger.info(f"Completion notifications: dispatching {len(claimed)} email(s)")
        sent = failed = 0
        with ThreadPoolExecutor(max_workers=_EMAIL_WORKERS) as pool:
            futures = {pool.submit(_send_completion_email, c): c for c in claimed}
            for future in as_completed(futures):
                c = futures[future]
                try:
                    ok = future.result()
                except Exception as e:
                    logger.error(f"Completion email error for {c['email']}: {e}")
                    ok = False
                # Success is recorded so it is never re-sent; failure releases the
                # claim so the next sync tries again.
                db.settle_notification(c['assignment_id'], c['email'], ok)
                if ok:
                    sent += 1
                else:
                    failed += 1

        logger.info(f"Completion notifications: {sent} sent, {failed} failed")
        return {'sent': sent, 'failed': failed, 'skipped': len(candidates) - len(claimed)}
    except Exception as e:
        logger.error(f"Error dispatching completion notifications: {e}")
        return {'sent': 0, 'failed': 0, 'skipped': 0}
    finally:
        _notify_lock.release()


def _send_completion_email(c):
    return send_course_completion_email(
        user_email=c['email'], user_name=c['name'], course_name=c['course'],
        completion_date=c['completion_date'], deadline=c.get('deadline'))


def fetch_fresh_data_from_api():
    """Fetch fresh data from API (blocking call)"""
    logger.info("Fetching fresh data from API (this may take 3+ minutes)...")
    response = requests.get(API_URL, timeout=API_TIMEOUT)
    response.raise_for_status()
    try:
        data = response.json()
    except ValueError:
        # Body was empty or not JSON (e.g. an HTML login/error page returned
        # with a 200 status). Surface what we actually received so the real
        # problem (expired token, redirect, wrong URL) is diagnosable.
        body_preview = (response.text or '').strip()[:500]
        logger.error(
            "API did not return JSON. status=%s content-type=%s final_url=%s body[:500]=%r",
            response.status_code,
            response.headers.get('Content-Type'),
            response.url,
            body_preview,
        )
        raise
    logger.info(f"Successfully fetched {len(data)} records from API")
    return data

def refresh_cache_background():
    """Refresh cache in background (called by scheduler)"""
    global _data_cache, _cache_timestamp, _refresh_in_progress
    
    with _cache_lock:
        if _refresh_in_progress:
            logger.info("Cache refresh already in progress, skipping...")
            return
        _refresh_in_progress = True
    
    try:
        logger.info("Background cache refresh started...")
        data = fetch_fresh_data_from_api()
        now = datetime.now()

        with _cache_lock:
            _data_cache = data
            _cache_timestamp = now
            save_cache_to_file(data, now)
        # Harvest this snapshot's completions into the append-only ledger so
        # prior-cycle dates survive the API overwriting its single latest value.
        record_completion_history(data, now)
        # Then congratulate anyone who has newly finished a course assigned to them.
        # Best-effort: an email problem must never fail the refresh.
        try:
            dispatch_completion_notifications(data)
        except Exception as e:
            logger.error(f"Completion notification pass failed: {e}")

        logger.info(f"Background cache refresh completed at {now.strftime('%Y-%m-%d %H:%M:%S')}")
    except Exception as e:
        logger.error(f"Error in background cache refresh: {e}")
    finally:
        with _cache_lock:
            _refresh_in_progress = False

def cache_refresh_scheduler():
    """Background thread that refreshes cache every N minutes.
    The API fetch itself is dispatched into a separate thread so the sleep
    interval is never shifted by a slow API response.
    """
    logger.info(f"Cache refresh scheduler started (interval: {AUTO_REFRESH_INTERVAL_MINUTES} minutes)")

    while True:
        time.sleep(AUTO_REFRESH_INTERVAL_MINUTES * 60)
        if not USE_LOCAL_DATA:
            try:
                t = threading.Thread(target=refresh_cache_background, daemon=True)
                t.start()
            except Exception as e:
                logger.error(f"Error in cache scheduler: {e}")

def initialize_cache():
    """Initialize cache on server startup"""
    global _data_cache, _cache_timestamp
    
    logger.info("Initializing cache system...")
    
    if USE_LOCAL_DATA:
        logger.info("Using local data mode - cache system disabled")
        return
    
    # Try to load from file first
    data, timestamp = load_cache_from_file()
    
    if data and timestamp:
        with _cache_lock:
            _data_cache = data
            _cache_timestamp = timestamp
        # Capture the loaded snapshot's completions in case the app was down when
        # the API last changed (idempotent — dedup'd by the ledger's primary key).
        record_completion_history(data, timestamp)
        # Catch up on completions that landed while the app was down, and baseline
        # any assignment not yet watched. Off the startup path so a slow SMTP server
        # can't delay the server coming up.
        threading.Thread(target=dispatch_completion_notifications, args=(data,),
                         daemon=True).start()

        # Check if cache is too old
        age_minutes = (datetime.now() - timestamp).total_seconds() / 60
        if age_minutes > AUTO_REFRESH_INTERVAL_MINUTES:
            logger.info(f"Cache is {age_minutes:.1f} minutes old, triggering refresh...")
            threading.Thread(target=refresh_cache_background, daemon=True).start()
        else:
            logger.info(f"Cache is {age_minutes:.1f} minutes old, still valid")
    else:
        logger.info("No valid cache found, fetching initial data...")
        threading.Thread(target=refresh_cache_background, daemon=True).start()
    
    # Start background scheduler
    scheduler_thread = threading.Thread(target=cache_refresh_scheduler, daemon=True)
    scheduler_thread.start()
    logger.info("Cache system initialized successfully")

def load_data():
    """Load data from cache (instant) or local file"""
    global _data_cache, _cache_timestamp
    
    if USE_LOCAL_DATA:
        # Load from local JSON file
        json_path = os.path.join(os.path.dirname(__file__), 'Response Sample.json')
        with open(json_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    else:
        # Return cached data (should always be available after initialization)
        with _cache_lock:
            if _data_cache is None:
                # Fallback: try to load from file if cache not initialized
                data, timestamp = load_cache_from_file()
                if data:
                    _data_cache = data
                    _cache_timestamp = timestamp
                else:
                    # Last resort: fetch synchronously (will block)
                    logger.warning("No cache available, fetching synchronously...")
                    data = fetch_fresh_data_from_api()
                    now = datetime.now()
                    _data_cache = data
                    _cache_timestamp = now
                    save_cache_to_file(data, now)
            
            return _data_cache

@app.route('/api/data')
@superadmin_required
def get_data():
    try:
        data = load_data()
        return jsonify(data)
    except Exception:
        import traceback
        logger.error(f"Error in get_data: {traceback.format_exc()}")
        return jsonify({"error": "Failed to load data"}), 500

@app.route('/api/raw-data')
@superadmin_required
def get_raw_data():
    try:
        data = load_data()
        return jsonify({'data': data})
    except Exception as e:
        logger.error(f"Error loading raw data: {e}")
        return jsonify({'error': 'Failed to load data'}), 500

@app.route('/api/summary')
@superadmin_required
def get_summary():
    try:
        print(f"Loading data (USE_LOCAL_DATA={USE_LOCAL_DATA})")
        data = load_data()
        print(f"Data received: {len(data)} records")
        
        # Process data for summary
        courses = defaultdict(lambda: {
            'name': '',
            'users': set(),
            'completed': set(),
            'in_progress': set(),
            'not_started': set(),
            'completion_dates': []
        })
        
        users = defaultdict(lambda: {
            'name': '',
            'email': '',
            'courses': set(),
            'completed_courses': set()
        })
        
        for record in data:
            course_name = record.get('Activity_Name', record.get('Course_Name', ''))
            user_email = record.get('User_Mail_ID', '')
            user_name = record.get('Participant_Name', '')
            completion_status = record.get('Course_Completion_Status', '')
            completion_date = record.get('Course_Completion_Date_(YYYY-MM-DD)', '').strip()
            
            # Course aggregation
            courses[course_name]['name'] = course_name
            courses[course_name]['users'].add(user_email)
            
            if completion_status == 'Completed':
                courses[course_name]['completed'].add(user_email)
                if completion_date:
                    courses[course_name]['completion_dates'].append(completion_date)
            elif completion_status == 'Current':
                activity_status = record.get('Activity_Status', '')
                if activity_status == 'Not Attempted':
                    courses[course_name]['not_started'].add(user_email)
                else:
                    courses[course_name]['in_progress'].add(user_email)
            
            # User aggregation
            users[user_email]['name'] = user_name
            users[user_email]['email'] = user_email
            users[user_email]['courses'].add(course_name)
            if completion_status == 'Completed':
                users[user_email]['completed_courses'].add(course_name)
        
        # Convert to list format
        course_list = []
        for course_name, course_data in courses.items():
            total_users = len(course_data['users'])
            completed = len(course_data['completed'])
            in_progress = len(course_data['in_progress'])
            not_started = len(course_data['not_started'])
            completion_rate = (completed / total_users * 100) if total_users > 0 else 0
            
            # Get earliest date
            earliest_date = None
            if course_data['completion_dates']:
                dates = [d for d in course_data['completion_dates'] if d]
                if dates:
                    earliest_date = min(dates)
            
            course_list.append({
                'course_name': course_name,
                'total_users': total_users,
                'completed': completed,
                'in_progress': in_progress,
                'not_started': not_started,
                'completion_rate': round(completion_rate, 2),
                'earliest_date': earliest_date
            })
        
        # Sort by total users descending
        course_list.sort(key=lambda x: x['total_users'], reverse=True)
        
        # User list
        user_list = []
        for user_email, user_data in users.items():
            total_courses = len(user_data['courses'])
            completed_courses = len(user_data['completed_courses'])
            completion_rate = (completed_courses / total_courses * 100) if total_courses > 0 else 0
            
            user_list.append({
                'user_name': user_data['name'],
                'user_email': user_email,
                'total_courses': total_courses,
                'completed_courses': completed_courses,
                'in_progress': total_courses - completed_courses,
                'completion_rate': round(completion_rate, 2)
            })
        
        # Sort by total courses descending
        user_list.sort(key=lambda x: x['total_courses'], reverse=True)
        
        # Calculate KPIs
        total_courses = len(courses)
        total_users = len(users)
        total_enrollments = sum(c['total_users'] for c in course_list)
        total_completions = sum(c['completed'] for c in course_list)
        overall_completion_rate = (total_completions / total_enrollments * 100) if total_enrollments > 0 else 0
        
        return jsonify({
            'kpis': {
                'total_courses': total_courses,
                'total_users': total_users,
                'total_enrollments': total_enrollments,
                'overall_completion_rate': round(overall_completion_rate, 2)
            },
            'courses': course_list,
            'users': user_list
        })
    except Exception:
        import traceback
        logger.error(f"Error in get_summary: {traceback.format_exc()}")
        return jsonify({"error": "Failed to load summary"}), 500

@app.route('/api/course/<path:course_name>')
@superadmin_required
def get_course_details(course_name):
    try:
        data = load_data()
        
        # Filter records for this course (using Activity_Name)
        course_records = [r for r in data if r.get('Activity_Name', r.get('Course_Name')) == course_name]
        
        # Group by user
        user_progress = {}
        for record in course_records:
            user_email = record.get('User_Mail_ID', '')
            if user_email not in user_progress:
                user_progress[user_email] = {
                    'user_name': record.get('Participant_Name', ''),
                    'user_email': user_email,
                    'completion_percentage': record.get('Course_Completion_Percentage', '0'),
                    'completion_status': record.get('Course_Completion_Status', ''),
                    'activity_status': record.get('Activity_Status', ''),
                     'completion_date': record.get('Course_Completion_Date_(YYYY-MM-DD)', '').strip()
                }
        
        return jsonify({
            'course_name': course_name,
            'users': list(user_progress.values())
        })
    except Exception:
        import traceback
        logger.error(f"Error in get_course_details: {traceback.format_exc()}")
        return jsonify({"error": "Failed to load course details"}), 500

@app.route('/api/user/<path:user_email>')
@superadmin_required
def get_user_details(user_email):
    try:
        data = load_data()
        
        # Filter records for this user
        user_records = [r for r in data if r.get('User_Mail_ID') == user_email]
        
        # Group by course
        course_progress = {}
        user_name = ''
        for record in user_records:
            course_name = record.get('Activity_Name', record.get('Course_Name', ''))
            user_name = record.get('Participant_Name', '')
            completion_date = record.get('Course_Completion_Date_(YYYY-MM-DD)', '').strip()
            
            if course_name not in course_progress:
                course_progress[course_name] = {
                    'course_name': course_name,
                    'completion_percentage': record.get('Course_Completion_Percentage', '0'),
                    'completion_status': record.get('Course_Completion_Status', ''),
                    'activity_status': record.get('Activity_Status', ''),
                    'completion_date': completion_date
                }
        
        # Sort by completion date
        courses_list = list(course_progress.values())
        courses_list.sort(key=lambda x: x['completion_date'] if x['completion_date'] else '9999-99-99')
        
        return jsonify({
            'user_name': user_name,
            'user_email': user_email,
            'courses': courses_list
        })
    except Exception:
        import traceback
        logger.error(f"Error in get_user_details: {traceback.format_exc()}")
        return jsonify({"error": "Failed to load user details"}), 500

# ── FY / Assignment Analytics dashboard ─────────────────────────────────────
# A second dashboard that analyses progress by *assignment* and *financial year*
# (see assignment_analytics.py). Unlike /api/summary (raw API), completion % here
# is measured against everyone assigned, and completions predating the assignment
# are treated as stale. Read-only analytics — open to any logged-in user.

def _load_assignments():
    return load_json_file(ASSIGNMENTS_FILE, {'assignments': []}).get('assignments', [])


def _include_untracked_arg():
    """?include_untracked=false hides users who are in the API but not on the uploaded
    roster. Anything other than an explicit 'false'/'0'/'no' means include them."""
    raw = (request.args.get('include_untracked') or 'true').strip().lower()
    return raw not in ('false', '0', 'no')


def _dimension_filters_arg():
    """?department=&location=&job_role= — the three roster dropdowns. Absent or 'all'
    means unconstrained; the NA sentinel selects users with nothing on record for that
    field. The engine validates the values against the options it built, so an unknown
    one is ignored rather than trusted."""
    return {dim: request.args.get(dim) for dim in aa.DIMENSIONS}


@app.route('/assignments-dashboard')
@login_required
def assignments_dashboard():
    return render_template('assignments.html', username=session.get('username'),
                           role=session.get('role'))


@app.route('/api/assignments-summary')
@login_required
def api_assignments_summary():
    """Full FY dashboard payload for a scope.

    ?fy=current|all|<start_year>  — financial year
    ?quarter=1..4|all             — narrow an FY scope to one of its quarters
    ?include_untracked=true|false — count users who are in the API but not on the
                                    uploaded roster
    ?department= ?location= ?job_role=
                                  — narrow to one roster department / location / role
                                    ('all' or absent = no constraint)
    """
    try:
        fy = request.args.get('fy', 'current')
        summary = aa.build_summary(
            _load_assignments(), load_data(), selected=fy,
            history_index=_history_index(),
            quarter=request.args.get('quarter'),
            profiles=_user_profiles(),
            include_untracked=_include_untracked_arg(),
            filters=_dimension_filters_arg())
        return jsonify(summary)
    except Exception:
        import traceback
        logger.error(f"Error in api_assignments_summary: {traceback.format_exc()}")
        return jsonify({'error': 'Failed to build assignments summary'}), 500


@app.route('/api/assignment-progress/<int:assignment_id>')
@login_required
def api_assignment_progress(assignment_id):
    """Per-user progress for one assignment (drill-down modal), with stale flags."""
    try:
        assignment = next((a for a in _load_assignments() if a.get('id') == assignment_id), None)
        if not assignment:
            return jsonify({'error': 'Assignment not found'}), 404
        index = aa.build_api_index(load_data())
        profiles = _user_profiles()
        prog = aa.compute_assignment_progress(assignment, index, _history_index())
        # Same population as the dashboard that opened this modal, so the modal's
        # totals always match the row that was clicked.
        if not _include_untracked_arg():
            prog = aa.drop_untracked(prog, profiles)
        # Annotate each row with its roster profile for the modal's extra columns.
        for u in prog['users']:
            p = profiles.get(u['email'].lower(), {})
            u['department'] = p.get('department', '')
            u['location'] = p.get('location', '')
            u['job_role'] = p.get('job_role', '')
            u['tracked'] = bool(p.get('tracked'))
            if p.get('name'):
                u['name'] = p['name']
        return jsonify(prog)
    except Exception:
        import traceback
        logger.error(f"Error in api_assignment_progress: {traceback.format_exc()}")
        return jsonify({'error': 'Failed to load assignment progress'}), 500


@app.route('/api/fy-user/<path:user_email>')
@login_required
def api_fy_user(user_email):
    """A single user's assignment-by-assignment breakdown within an FY scope."""
    try:
        fy = request.args.get('fy', 'current')
        quarter = aa.parse_quarter(request.args.get('quarter'))
        assignments = _load_assignments()
        index = aa.build_api_index(load_data())
        history_index = _history_index()
        resolved = aa.resolve_selected_fy(fy, assignments)

        rows = []
        user_name = ''
        for a in assignments:
            p = aa.compute_assignment_progress(a, index, history_index)
            if resolved != 'all' and p['fy_start_year'] != resolved:
                continue
            if resolved != 'all' and quarter and p['quarter'] != quarter:
                continue
            u = next((x for x in p['users'] if x['email'] == user_email), None)
            if not u:
                continue
            if u['name'] and not user_name:
                user_name = u['name']
            rows.append({
                'assignment_id': a.get('id'),
                'title': p['title'],
                'course_name': p['course_name'],
                'created_date': p['created_date'],
                'fy_label': p['fy_label'],
                'deadline': p['deadline'],
                'status': u['status'],
                'percentage': u['percentage'],
                'completion_date': u['completion_date'],
                'stale': u['stale'],
                'untouched': u['untouched'],
            })
        rows.sort(key=lambda r: r['created_date'], reverse=True)
        return jsonify({'user_email': user_email, 'user_name': user_name, 'assignments': rows})
    except Exception:
        import traceback
        logger.error(f"Error in api_fy_user: {traceback.format_exc()}")
        return jsonify({'error': 'Failed to load user assignment details'}), 500


# Initialize cache when the module is imported (covers both direct run and
# Waitress/any WSGI server that imports app without running __main__).
initialize_cache()

if __name__ == '__main__':
    # use_reloader=False is required: with the default reloader Flask spawns a
    # child subprocess for requests and a parent watcher process. The scheduler
    # would start twice — once in the parent watcher and once in the child.
    app.run(debug=True, port=5000, use_reloader=False)
